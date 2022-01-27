import openpyxl

def getRowCount(file, sheetName):
   wb=openpyxl.load_workbook(file)
   sheet=wb.get_sheet_by_name(sheetName)
   return (sheet.max_row)

def getColumnCount(file, sheetName):
   wb=openpyxl.load_workbook(file)
   sheet=wb.get_sheet_by_name(sheetName)
   return (sheet.max_column)

def readData(file, sheetName, rownum, columnnum):
   wb=openpyxl.load_workbook(file)
   sheet=wb.get_sheet_by_name(sheetName)
   return sheet.cell(row=rownum, column=columnnum).value

def writeData(file, sheetName, rownum, columnnum, data):
   wb=openpyxl.load_workbook(file)
   sheet=wb.get_sheet_by_name(sheetName)
   sheet.cell(row=rownum, column=columnnum).value=data
   wb.save(file)
